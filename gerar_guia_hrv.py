"""
Guia de Compra — Honda HR-V EXL 1.8 2017/2018
Gera planilha Excel com agenda de revisões, perguntas ao vendedor e checklist de visita.
Execute: python gerar_guia_hrv.py
"""

import os
import webbrowser

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Instalando openpyxl...")
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def borda():
    s = Side(border_style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


def cab(ws, row, col, valor, bg="2563EB", fg="FFFFFF", bold=True, size=11,
        h_align="center", wrap=False):
    c = ws.cell(row=row, column=col, value=valor)
    c.font = Font(bold=bold, size=size, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    c.border = borda()
    return c


def cel(ws, row, col, valor, bg="FFFFFF", bold=False, size=11,
        h_align="left", wrap=True, cor_fonte="000000"):
    c = ws.cell(row=row, column=col, value=valor)
    c.font = Font(bold=bold, size=size, color=cor_fonte)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=h_align, vertical="top", wrap_text=wrap)
    c.border = borda()
    return c


# ═══════════════════════════════════════════════════════════════════════════════
#  ABA 1 — MANUTENÇÕES
# ═══════════════════════════════════════════════════════════════════════════════

MANUTENCOES = [
    # (item, descrição, km, tempo, custo_estimado, critico, pergunta_vendedor)
    (
        "Óleo do motor",
        "Troca do óleo sintético 5W30 (Honda Genuine Oil ou equivalente API SN/SP). "
        "O motor 1.8 i-VTEC consome aprox. 4,2 L. Nunca usar óleo mineral.",
        "10.000 km", "12 meses", "R$ 200–350",
        "SIM",
        "Tem notas fiscais das trocas de óleo? Qual a marca/viscosidade usada? "
        "Quando foi a última troca e quantos km tem hoje?"
    ),
    (
        "Filtro de óleo",
        "Trocado sempre junto com o óleo. Use filtro original Honda ou Tecfil/WIX. "
        "Nunca reaproveite.",
        "10.000 km", "12 meses", "R$ 30–60",
        "Não",
        "Foi trocado na última revisão junto com o óleo?"
    ),
    (
        "Filtro de ar do motor",
        "Filtro de admissão de ar. Afeta consumo, desempenho e durabilidade do motor. "
        "Em regiões com muito pó (interior), trocar antes do prazo.",
        "30.000 km", "24 meses", "R$ 60–120",
        "Não",
        "Quando foi trocado pela última vez? Posso ver o filtro agora?"
    ),
    (
        "Filtro do ar-condicionado (cabine)",
        "Filtro do habitáculo. Influencia na qualidade do ar, eficiência do A/C e "
        "cheiro dentro do carro. Acesso simples — fica atrás do porta-luvas.",
        "15.000 km", "12 meses", "R$ 40–80",
        "Não",
        "O A/C tem algum cheiro estranho quando liga? O filtro foi trocado recentemente?"
    ),
    (
        "Fluido do câmbio CVT  ⚠️ MAIS CRÍTICO",
        "PONTO MAIS CRÍTICO DO HR-V! O câmbio CVT exige fluido exclusivo Honda HCF-2 "
        "(NÃO usar ATF genérico). Sem manutenção regular, o CVT pode falhar antes de "
        "100.000 km. Reparo/troca do CVT custa R$ 5.000–15.000.",
        "40.000 km", "—", "R$ 300–500",
        "SIM ⚠️",
        "O câmbio CVT já foi revisado? Tem nota fiscal de troca de fluido? "
        "O câmbio patina em subida? Faz solavanco ao sair do lugar? "
        "Tem qualquer barulho ou comportamento estranho?"
    ),
    (
        "Velas de ignição",
        "HR-V 1.8 usa velas iridium NGK ILZKR7B-11S (4 unidades). "
        "Iridium tem vida longa mas quando gastam aumentam consumo e engasgam o motor.",
        "50.000 km", "—", "R$ 200–350",
        "Não",
        "Já foram trocadas? O motor tem algum engasgamento ou demora para ligar a frio?"
    ),
    (
        "Líquido de freio (DOT 3)",
        "Fluido de freio absorve umidade ao longo do tempo, reduzindo o ponto de "
        "ebulição e eficiência da frenagem. Troca simples e barata.",
        "40.000 km", "2 anos", "R$ 80–150",
        "Não",
        "Quando foi trocado pela última vez? O pedal de freio fica firme ou esponjoso?"
    ),
    (
        "Pastilhas de freio dianteiras",
        "HR-V pesa ~1.200 kg com freio a disco nas 4 rodas (EXL). "
        "Pastilhas dianteiras desgastam mais rápido. Use Cobreq, Ferodo ou original Honda.",
        "30.000–40.000 km", "—", "R$ 150–350",
        "Não",
        "Quando foram trocadas? O freio faz algum barulho (rangido, apito ao frear)?"
    ),
    (
        "Pastilhas de freio traseiras",
        "Freio traseiro a disco — desgaste menor que o dianteiro mas deve ser verificado.",
        "40.000–60.000 km", "—", "R$ 120–280",
        "Não",
        "Foram trocadas alguma vez? O carro puxa para um lado ao frear forte?"
    ),
    (
        "Fluido de arrefecimento (radiador)",
        "Honda Long Life Coolant (azul). O sistema é fechado — se precisar repor "
        "frequentemente, investigar vazamento. Cor turva indica degradação.",
        "120.000 km", "5 anos", "R$ 100–200",
        "Não",
        "A temperatura do motor fica sempre no meio do marcador? "
        "Alguma vez a luz de temperatura acendeu no painel?"
    ),
    (
        "Corrente de distribuição",
        "VANTAGEM: HR-V 1.8 usa CORRENTE (não correia)! Não precisa trocar. "
        "Apenas verificar se há barulho metálico em motor frio (tensor desgastado).",
        "Não troca", "—", "—",
        "INFO ✓",
        "O motor faz algum barulho metálico nos primeiros segundos após ligar? "
        "(Deve ser silencioso — ruído = tensor do tensor da corrente desgastado)"
    ),
    (
        "Pneus (215/60 R16)",
        "Verificar: desgaste uniforme (câmara ok), profundidade mínima 1,6mm "
        "(ideal >4mm), data de fabricação no DOT (máx. 5 anos), calibragem 32 PSI. "
        "Marcas ok: Michelin, Bridgestone, Continental, Pirelli.",
        "Rodízio: 10.000 km", "Calibragem: mensal", "R$ 400–800 (jogo)",
        "SIM",
        "Qual a marca e data de fabricação dos pneus (número DOT)? "
        "O carro puxa para algum lado? O volante treme acima de 80 km/h?"
    ),
    (
        "Alinhamento e balanceamento",
        "Essencial para segurança, desgaste uniforme dos pneus e conforto. "
        "Fazer sempre após troca de pneu, batida ou a cada 10.000 km.",
        "10.000 km", "6 meses", "R$ 100–180",
        "Não",
        "Quando foi feito? O volante fica reto em reta? O carro desvia sem segurar o volante?"
    ),
    (
        "Amortecedores",
        "Verificar vazamento de óleo e desgaste. Amortecedor ruim compromete "
        "segurança (maior distância de frenagem) e conforto.",
        "80.000 km", "—", "R$ 600–1.200 (par)",
        "Não",
        "O carro bate muito em buracos? Balança excessivamente após uma lombada? "
        "Já foram trocados?"
    ),
    (
        "Bateria (45 Ah)",
        "Vida útil média 3–4 anos. Verificar data de fabricação (impressa na bateria). "
        "Bateria fraca causa partida lenta e pode danificar o alternador.",
        "50.000 km", "3–4 anos", "R$ 250–450",
        "Não",
        "Qual o ano da bateria? A partida é rápida e firme mesmo em dias frios?"
    ),
    (
        "Ar-condicionado (gás R134a)",
        "O EXL tem A/C digital com controle de temperatura. "
        "Recarga de gás não é manutenção periódica — só fazer se perder pressão. "
        "Verificar se o compressor liga (barulho suave ao acionar).",
        "—", "Verificar anualmente", "R$ 150–300 (recarga)",
        "Não",
        "O A/C gela bem mesmo em dias de 35°C? Já foi recarregado? "
        "O compressor faz barulho quando liga o A/C?"
    ),
    (
        "Buchas da suspensão / bandejas",
        "Componentes de borracha que se degradam com o tempo. "
        "Sintoma: barulho de 'clique' ou 'estalo' em curvas, buracos ou ao frear.",
        "60.000–80.000 km", "—", "R$ 500–1.500",
        "Não",
        "O carro faz algum barulho de estalo em curvas ou ao passar por buracos? "
        "A direção tem folga ou vibração?"
    ),
    (
        "Revisão completa (concessionária Honda)",
        "Revisão periódica com substituição de todos os itens da agenda Honda. "
        "Registro na caderneta ou sistema Honda é prova de manutenção preventiva.",
        "30.000 km", "12 meses", "R$ 500–900",
        "SIM",
        "Tem caderneta de revisões ou histórico no sistema Honda? "
        "Quantas revisões foram feitas na concessionária? "
        "Pode me mostrar as notas fiscais ou comprovantes?"
    ),
]


def criar_aba_manutencao(wb):
    ws = wb.active
    ws.title = "Manutenções"

    # Título
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "HONDA HR-V EXL 1.8  2017/2018  —  Agenda Completa de Revisões e Manutenção"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1E3A5F")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    hdrs = ["#", "Item", "Descrição Técnica", "A cada (km)",
            "A cada (tempo)", "Custo estimado", "Crítico?", "Pergunta ao Vendedor"]
    for col, h in enumerate(hdrs, 1):
        cab(ws, 2, col, h, bg="1D4ED8")
    ws.row_dimensions[2].height = 24

    CORES = {"SIM": "FEE2E2", "SIM ⚠️": "FEE2E2",
             "INFO ✓": "DCFCE7", "Não": None}

    for i, (item, desc, km, tempo, custo, critico, pergunta) in enumerate(MANUTENCOES, 1):
        row = i + 2
        bg = CORES.get(critico, None)
        # alterna cinza/branco quando não crítico
        if bg is None:
            bg = "F9FAFB" if i % 2 == 0 else "FFFFFF"

        cel(ws, row, 1, i, bg=bg, bold=True, h_align="center", wrap=False)
        cel(ws, row, 2, item, bg=bg, bold=True)
        cel(ws, row, 3, desc, bg=bg)
        cel(ws, row, 4, km, bg=bg, h_align="center", wrap=False)
        cel(ws, row, 5, tempo, bg=bg, h_align="center", wrap=False)
        cel(ws, row, 6, custo, bg=bg, h_align="center", wrap=False)

        cor_critico = "DC2626" if "SIM" in critico else ("166534" if "INFO" in critico else "374151")
        cel(ws, row, 7, critico, bg=bg, bold=True, h_align="center", wrap=False,
            cor_fonte=cor_critico)
        cel(ws, row, 8, pergunta, bg=bg)
        ws.row_dimensions[row].height = 70

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 17
    ws.column_dimensions["G"].width = 11
    ws.column_dimensions["H"].width = 52
    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════════════
#  ABA 2 — PERGUNTAS AO VENDEDOR
# ═══════════════════════════════════════════════════════════════════════════════

PERGUNTAS = [
    ("DOCUMENTAÇÃO", [
        ("O carro está no seu nome? Pode mostrar o CRLV?",
         "Evita surpresas com transferências pendentes, bloqueios judiciais ou penhoras"),
        ("IPVA e licenciamento estão em dia?",
         "IPVA atrasado gera multa de 20% + juros. Licenciamento vencido = multa de R$ 293"),
        ("Tem débitos de multas no RENAVAM?",
         "Consulte em detran.sp.gov.br antes de fechar. Débitos passam para o comprador"),
        ("Aceita fazer um laudo cautelar antes de fechar?",
         "Laudo custa ~R$ 200–400 e detecta adulteração de chassi, recuperação de acidente, "
         "pintura original vs repintura — essencial para qualquer compra acima de R$ 50k"),
        ("O veículo tem financiamento em aberto ou restrição judicial?",
         "Consulte no banco emissor do CRV e no site do DETRAN-SP com o RENAVAM"),
    ]),
    ("HISTÓRICO", [
        ("É único dono? Desde quando o veículo está com você?",
         "Único dono costuma ter melhor cuidado e histórico documentado"),
        ("Por que está vendendo?",
         "Resposta evasiva é sinal de alerta. Razões legítimas: troca de modelo, "
         "mudança de cidade, comprou 0km, carro sobrando na família"),
        ("O carro já sofreu acidente? Qual foi o impacto?",
         "Acidentes grandes comprometem a estrutura mesmo após reparos. "
         "Pergunte sobre funilaria, troca de peças e pintura"),
        ("Tem caderneta de revisões ou notas fiscais dos serviços?",
         "Histórico documentado vale dinheiro — especialmente para o câmbio CVT"),
        ("O carro ficou parado por algum período longo (mais de 3 meses)?",
         "Carro parado deteriora borrachas, fluidos, freios e bateria. "
         "Risco de ferrugem interna em peças móveis"),
    ]),
    ("MECÂNICA — PONTOS CRÍTICOS DO HR-V", [
        ("O câmbio CVT já foi revisado? Tem nota fiscal de troca de fluido?",
         "PONTO MAIS CRÍTICO! CVT sem manutenção pode falhar antes de 100.000 km. "
         "Reparo custa R$ 5.000–15.000. Insista em ver comprovante"),
        ("O câmbio patina em subida, solavanca ao sair parado ou faz barulho?",
         "Sintomas de CVT com problema: patinação em subida, solavanco ao sair do lugar, "
         "barulho de 'correia' ou 'chiado' em aceleração"),
        ("O motor faz barulho metálico nos primeiros segundos após ligar a frio?",
         "HR-V usa corrente (não correia) — normalmente silencioso. "
         "Barulho metálico pode indicar tensor da corrente desgastado (~R$ 1.500)"),
        ("O A/C gela bem mesmo em dias de muito calor (acima de 32°C)?",
         "Compressor fraco ou com vazamento = A/C que não gela. "
         "Recarga de gás é barata mas compressor novo pode custar R$ 1.500–3.000"),
        ("Os freios têm pedal firme? Fazem algum barulho ao frear?",
         "Rangido = pastilha gasta (barata de resolver). "
         "Pedal mole/esponjoso = fluido de freio degradado ou problema no sistema hidráulico"),
        ("Os pneus têm desgaste uniforme nos dois lados?",
         "Desgaste irregular indica problema de alinhamento ou suspensão. "
         "Verifique a data de fabricação no código DOT na lateral do pneu"),
    ]),
    ("NEGOCIAÇÃO  (dicas do especialista)", [
        ("Pesquise o preço FIPE antes — HR-V EXL 2017 está ~R$ 83.000 (FIPE 2026)",
         "Se o preço pedido for muito acima da FIPE, questione: 'A FIPE está em R$ X — "
         "como você chegou nesse preço?'"),
        ("Use defeitos encontrados para pedir desconto concreto",
         "Pneus gastos: -R$ 600. Pastilha gasta: -R$ 350. CVT sem histórico: -R$ 2.000. "
         "Pintura diferente em uma porta: -R$ 800. Seja específico nos valores"),
        ("Proponha pagamento à vista via PIX ou TED no fechamento",
         "À vista tem poder real. Pedir 5–8% de desconto é razoável e esperado"),
        ("Nunca demonstre pressa ou entusiasmo excessivo",
         "Diga que está visitando outros 3–4 carros essa semana. "
         "Isso é verdade se você usar o buscador — e aumenta seu poder de negociação"),
        ("Solicite 24h para dar a resposta final",
         "Use esse tempo para pesquisar o RENAVAM no DETRAN-SP e consultar um mecânico "
         "de confiança. Vendedor que não aceita 24h está pressionando — sinal de alerta"),
        ("Se tiver qualquer dúvida no câmbio CVT, peça R$ 2.000 de desconto ou desista",
         "CVT é o ponto mais caro de manutenção do HR-V. "
         "Sem histórico comprovado, assuma que precisará de revisão nos próximos 20.000 km"),
    ]),
]


def criar_aba_perguntas(wb):
    ws = wb.create_sheet("Perguntas ao Vendedor")

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "Roteiro Completo de Perguntas — Especialista em Honda HR-V EXL"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1E3A5F")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    row = 2
    for categoria, items in PERGUNTAS:
        ws.merge_cells(f"A{row}:C{row}")
        c = ws[f"A{row}"]
        c.value = f"  {categoria}"
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1D4ED8")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = borda()
        ws.row_dimensions[row].height = 22
        row += 1

        for pergunta, motivo in items:
            cel(ws, row, 1, "?", bg="EFF6FF", bold=True, h_align="center", wrap=False,
                cor_fonte="1D4ED8")
            cel(ws, row, 2, pergunta, bg="FFFFFF", bold=True)
            cel(ws, row, 3, motivo, bg="F0F9FF")
            ws.row_dimensions[row].height = 45
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 58
    ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════════════════════
#  ABA 3 — CHECKLIST NA VISITA
# ═══════════════════════════════════════════════════════════════════════════════

CHECKLIST = [
    ("EXTERIOR  (fazer antes de entrar no carro)", [
        ("Pintura uniforme — sem diferença de tom em nenhum painel",
         "Olhe de lado com sol ou lanterna do celular. Tom diferente = repintura = batida"),
        ("Sem amassados, mossas ou rachaduras em para-choques",
         "Percorra toda a carroceria devagar com a mão — mossas pequenas são difíceis de ver"),
        ("Frestas uniformes entre portas, capô e porta-mala",
         "Fresta desigual indica que a peça foi removida ou o carro sofreu batida estrutural"),
        ("Pneus sem desgaste irregular (desgaste só de um lado = suspensão com problema)",
         "Passe a mão na banda de rodagem. Desgaste irregular = alinhamento ou bucha ruim"),
        ("Luzes: faróis, lanternas, pisca-pisca e luz de freio funcionando",
         "Peça a alguém para pisar no freio enquanto você olha atrás"),
    ]),
    ("INTERIOR  (com o motor desligado)", [
        ("Sem cheiro de mofo, umidade ou cigarro forte",
         "Mofo = infiltração de água (teto solar, vidros, borrachas). Difícil de eliminar"),
        ("Painel sem luzes de advertência acesas após a ignição",
         "Ligue a chave e espere 10 segundos — todas as luzes devem apagar"),
        ("A/C resfria bem em todas as temperaturas e velocidades do ventilador",
         "Teste no nível máximo. EXL tem A/C digital — verifique o display"),
        ("Tela multimídia funciona, conecta Bluetooth e mostra câmera de ré",
         "EXL tem câmera de ré — verifique se a imagem está nítida e sem pixelação"),
        ("Todos os 4 vidros elétricos sobem e descem completamente",
         "Vidro que para no meio = motor elétrico fraco (~R$ 300–500 para trocar)"),
        ("Travas elétricas funcionam nas 4 portas",
         "Trave e destrave cada porta individualmente"),
        ("Banco de couro sem rasgos, costuras soltando ou manchas",
         "EXL tem couro — costuras que soltam são difíceis de reparar"),
        ("Air bag: luz apaga após ligar o carro",
         "Luz de air bag acesa = bag pode ter sido acionado e rearmado. Fuja!"),
    ]),
    ("MECÂNICA  (test drive obrigatório)", [
        ("Motor silencioso ao ligar a frio — sem barulho metálico",
         "Ligue sem aquecer e ouça por 30 segundos. Deve ser quase silencioso"),
        ("Sem fumaça azul no escapamento (=queima óleo) ou preta (=combustão rica)",
         "Olhe o escapamento ao ligar e durante aceleração"),
        ("CVT faz transição suave — sem solavanco ao sair do lugar ou patinação em subida",
         "Saia do lugar em subida. CVT bom é imperceptível. Patinação = problema grave"),
        ("Sem desvio de rota ao frear forte de ~60 km/h",
         "Carro que puxa ao frear = pastilha desgastada de um lado ou disco empenado"),
        ("Freio de estacionamento segura o carro em rampa",
         "Teste na primeira ladeira que encontrar"),
        ("Direção sem vibração no volante acima de 80 km/h",
         "Vibração = balanceamento ou cubo de roda com problema"),
        ("Sem barulho de estalo em curvas ou manobras",
         "Estalo em curva = bucha do estabilizador ou bandeja desgastada"),
        ("Sem ruídos, batidas ou vibrações no piso no geral",
         "Barulho no piso = amortecedor, calço, coxim ou solda da carroceria"),
    ]),
    ("DOCUMENTAÇÃO  (verificar antes de assinar)", [
        ("Número do chassi (VIN) no painel confere com o CRLV",
         "Chassi fica visível pelo para-brisa, lado motorista. Rasuras ou troca = fraude"),
        ("CPF/CNPJ no CRLV é o mesmo do vendedor",
         "Se o nome for diferente, exija a cadeia completa de transferências"),
        ("Laudo cautelar aprovado por vistoriador credenciado DETRAN",
         "Não dispense o laudo. Ele detecta o que o olho não vê"),
        ("Pesquisa no DETRAN-SP: sem multas, sem restrições, IPVA quitado",
         "Acesse detran.sp.gov.br/veiculo com o RENAVAM. Feito em 2 minutos"),
    ]),
]


def criar_aba_checklist(wb):
    ws = wb.create_sheet("Checklist na Visita")

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Checklist na Hora da Visita — Honda HR-V EXL 2017  (imprima e leve)"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="166534")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    hdrs = ["OK?", "O que verificar", "Como fazer / O que significa", "Observação"]
    for col, h in enumerate(hdrs, 1):
        cab(ws, 2, col, h, bg="15803D")
    ws.row_dimensions[2].height = 22

    row = 3
    for categoria, items in CHECKLIST:
        ws.merge_cells(f"A{row}:D{row}")
        c = ws[f"A{row}"]
        c.value = f"  {categoria}"
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="374151")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = borda()
        ws.row_dimensions[row].height = 20
        row += 1

        for i, (item, como) in enumerate(items):
            bg = "F0FDF4" if i % 2 == 0 else "FFFFFF"
            cel(ws, row, 1, "☐", bg=bg, h_align="center", wrap=False,
                bold=True, cor_fonte="15803D")
            cel(ws, row, 2, item, bg=bg, bold=True)
            cel(ws, row, 3, como, bg=bg)
            cel(ws, row, 4, "", bg="FAFAFA")
            ws.row_dimensions[row].height = 38
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 22
    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  Guia de Compra — Honda HR-V EXL 1.8 2017/2018")
    print("=" * 60)

    wb = openpyxl.Workbook()
    criar_aba_manutencao(wb)
    criar_aba_perguntas(wb)
    criar_aba_checklist(wb)

    nome = "guia_compra_hrv.xlsx"
    wb.save(nome)
    caminho = os.path.abspath(nome)
    print(f"\n  Planilha gerada: {caminho}")
    print("  Abrindo...")
    webbrowser.open(caminho)


if __name__ == "__main__":
    main()
