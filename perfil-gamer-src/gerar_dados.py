# Gera ../perfil-gamer/dados.js a partir de biblioteca_jogos.xlsx
# Fluxo de atualização: substituir o xlsx -> rodar este script -> commit/PR/merge.
import openpyxl, json, datetime, os

os.chdir(os.path.dirname(os.path.abspath(__file__)))
wb = openpyxl.load_workbook('biblioteca_jogos.xlsx', data_only=True)
ws = wb['Biblioteca de Jogos']

games = []
for r in range(2, ws.max_row + 1):
    nome = ws.cell(r, 2).value
    if not nome:
        continue
    games.append({
        'id': ws.cell(r, 1).value,
        'nome': str(nome),
        'horas': ws.cell(r, 3).value or 0,
        'enc': ws.cell(r, 4).value,          # S / N / #
        'w': ws.cell(r, 5).value or 1,
        'motivo': ws.cell(r, 6).value or '',
        'genres': ws.cell(r, 7).value or '',
        'h': ws.cell(r, 8).value,
        'g': ws.cell(r, 9).value,
        'd': ws.cell(r, 10).value,
        'ff': ws.cell(r, 11).value,
        'v': ws.cell(r, 12).value or 'N/A',
        'obs': ws.cell(r, 13).value or ''
    })

# Planejamento 2026 (manter em dia manualmente aqui)
plano = [
    {"jogo": "007 First Light", "lanc": "Jun/26", "est": "15h", "status": "—"},
    {"jogo": "AC Black Flag Resync", "lanc": "09/Jul", "est": "45h", "status": "❌ Dropado com 13h (jul/26) — liberou ~32h"},
    {"jogo": "Beast of Reincarnation", "lanc": "05/Ago", "est": "45h", "status": "Próximo"},
    {"jogo": "Lord of the Fallen 2", "lanc": "10/Ago", "est": "40h", "status": "Sugerido p/ Setembro"},
    {"jogo": "Blood of Dawn", "lanc": "03/Set", "est": "45h", "status": "Outubro"},
    {"jogo": "Wolverine", "lanc": "15/Set", "est": "25h", "status": "Novembro"},
    {"jogo": "Phantom Blade", "lanc": "08/Set", "est": "50h", "status": "Nov–Dez"},
]
notas_plano = [
    "Disponibilidade: ~12h/semana (~52h/mês; setembro reduzido ~34h por férias 08–27/09)",
    "Em andamento (jul/26): AC Shadows NG+ nightmare (comprado R$60) · Exp33 rezerado (+17h)",
    "FF14: assinatura pausada; retorno out/nov 2026 antes da expansão Evercold (jan/2027)",
]

insights = [
    ["Gameplay é o motor do veredicto", "História amplifica, não sustenta; para souls-likes história não é eixo"],
    ["Fun Factor 4 → conclusão garantida", "Padrão histórico de 100%"],
    ["Horas jogadas ≠ qualidade", "Valhalla 131h = Medíocre (jogo gigante + pandemia); Tsushima 65h = Masterpiece"],
    ["Retentativa de jogo dropado: 0% de sucesso", "Nenhuma passou de Medíocre — diferente de replay de jogo zerado"],
    ["Walkthrough ≥2 = jogo amado (ou DLC)", "Tsushima e Exp33 Masterpiece, TLoU; exceção FFXV (DLCs)"],
    ["Fórmula Ubisoft = risco alto", "Odyssey/Origins/Outlaws Muito Ruim e drop; exceção parcial: Shadows em NG+"],
    ["Souls-like só FromSoftware", "Elden Ring/Sekiro Masterpiece vs LotF/Demon Souls Medíocre, Khazan/Wuchang drop"],
    ["Survival Horror: 100% conclusão", "Gênero subestimado na coleção"],
    ["Sem fast travel fácil = drop", "Dragon's Dogma 2, FFXII, Kingdom Come II"],
    ["Combate simples = drop mesmo com boa escrita", "AC Black Flag Resync: cutscenes boas, dropado em 13h"],
]

hoje = datetime.date.today().strftime('%d/%m/%Y')
out = (
    "// GERADO por perfil-gamer-src/gerar_dados.py — nao editar na mao\n"
    f"window.DADOS_ATUALIZADOS = {json.dumps(hoje)};\n"
    f"window.JOGOS = {json.dumps(games, ensure_ascii=False, indent=1)};\n"
    f"window.PLANO = {json.dumps(plano, ensure_ascii=False, indent=1)};\n"
    f"window.NOTAS_PLANO = {json.dumps(notas_plano, ensure_ascii=False, indent=1)};\n"
    f"window.INSIGHTS = {json.dumps(insights, ensure_ascii=False, indent=1)};\n"
)
open('../perfil-gamer/dados.js', 'w', encoding='utf-8').write(out)
print('dados.js gerado:', len(games), 'jogos')
