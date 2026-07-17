# Gera biblioteca_jogos.md a partir de biblioteca_jogos.xlsx
import openpyxl, json, datetime
from collections import Counter, defaultdict

VERSAO = 'v11'
HOJE = datetime.date.today().strftime('%d/%m/%Y')

wb = openpyxl.load_workbook('biblioteca_jogos.xlsx', data_only=True)
ws = wb['Biblioteca de Jogos']
num = json.load(open('.notas_numericas_backup.json', encoding='utf-8'))

ORD = ['Muito Ruim','Ruim','Medíocre','Bom','Muito Bom','Masterpiece']
EMO = {'Masterpiece':'🏆','Muito Bom':'🟢','Bom':'🔵','Medíocre':'🟡','Ruim':'🟠','Muito Ruim':'🔴'}
def show(v): return f'{EMO[v]} {v}' if v in EMO else 'N/A'

games = []
for r in range(2, ws.max_row+1):
    g = dict(id=ws.cell(r,1).value, nome=ws.cell(r,2).value, horas=ws.cell(r,3).value or 0,
             enc=ws.cell(r,4).value, w=ws.cell(r,5).value, motivo=ws.cell(r,6).value,
             genres=ws.cell(r,7).value or '', h=ws.cell(r,8).value, g=ws.cell(r,9).value,
             d=ws.cell(r,10).value, ff=ws.cell(r,11).value, v=ws.cell(r,12).value,
             obs=ws.cell(r,13).value or '')
    if g['nome']:
        g['num'] = num.get(g['nome'])
        games.append(g)

total=len(games); comp=[g for g in games if g['enc']=='S']; drop=[g for g in games if g['enc']=='N']
prog=[g for g in games if g['enc']=='#']; horas_tot=sum(g['horas'] for g in games)
rated=[g for g in games if g['v'] in ORD]
dist=Counter(g['v'] for g in rated)

def vmed(lst):
    ns=[g['num'] for g in lst if isinstance(g['num'],(int,float))]
    if not ns: return 'N/A'
    m=sum(ns)/len(ns)
    for lim,v in [(4.75,'Masterpiece'),(3.75,'Muito Bom'),(2.75,'Bom'),(2.25,'Medíocre'),(1.25,'Ruim')]:
        if m>=lim: return show(v)
    return show('Muito Ruim')

L=[]; A=L.append
A('# 🎮 Biblioteca de Jogos — PS4 + PS5\n')
A(f'> **Documentação {VERSAO}** · Atualizada em {HOJE}  ')
A('> **Sistema de Veredictos (estilo ACG):** Masterpiece · Muito Bom · Bom · Medíocre · Ruim · Muito Ruim  ')
A('> Medíocre = mediano, não ruim — não há nada de errado em ser mediano\n')
A('---\n')
A('## 🗂️ Esquema de Dados (base para a aplicação)\n')
A('Arquivo fonte: `biblioteca_jogos.xlsx` · Aba `Biblioteca de Jogos` · Gerador: `gerar_doc.py`\n')
A('| Coluna | Tipo | Descrição |')
A('|---|---|---|')
A('| # | int | ID sequencial |')
A('| Jogo | texto | Nome do jogo |')
A('| Tempo Total | horas | Horas totais (PS4+PS5 consolidado) |')
A('| Encerrado | S / N / # | S=Completou · N=Droppou · #=Em progresso |')
A('| Qtde Walkthrough | int | Nº de walkthroughs |')
A('| Motivo Desistência | texto | Apenas quando N |')
A('| Genres | texto | 1º gênero = principal |')
A('| História (1-4) | int ou N/A | Componente interno de análise |')
A('| Gameplay (1-4) | int | Componente interno de análise |')
A('| Dificuldade (1-3) | int | 1=Fácil · 2=Equilibrado · 3=Difícil |')
A('| Fun Factor (1-4) | int | Componente interno de análise |')
A('| Veredicto | enum 6 níveis | Masterpiece / Muito Bom / Bom / Medíocre / Ruim / Muito Ruim |')
A('| Observação | texto | Contexto livre |\n')
A('**Filosofia do veredicto:** julgamento categórico, não medição. Componentes 1–4 são evidência analítica interna.  ')
A('**Ordem ordinal (p/ app):** Muito Ruim=1 · Ruim=2 · Medíocre=3 · Bom=4 · Muito Bom=5 · Masterpiece=6  ')
A('**Parâmetro econômico fixo:** custo líquido médio por jogo = **R$90** (via grupo + revenda). Custo/hora = R$90 ÷ Tempo Total.  ')
A('**Agente de apoio:** ver `agente_perfil_gamer.md` — especialista que categoriza e analisa junto com o usuário.\n')
A('---\n')
A('## 📊 Resumo Geral\n')
A('| Métrica | Valor | Métrica | Valor |')
A('|---|---|---|---|')
A(f'| Total de jogos | **{total}** | Total de horas | **~{horas_tot:,.0f}h** |'.replace(',','.'))
A(f'| ✅ Completados | **{len(comp)}** ({len(comp)/total:.0%}) | Média horas/jogo | **{horas_tot/total:.0f}h** |')
A(f'| ❌ Droppados | **{len(drop)}** ({len(drop)/total:.0%}) | Veredicto mais comum | **{dist.most_common(1)[0][0]}** |')
A(f"| 🔄 Em progresso | **{len(prog)}** | Masterpieces | **{dist['Masterpiece']}** jogos |\n")
A('---\n')
A('## 📋 Lista Completa\n')
A('| # | Jogo | Genres | Total | Enc | W | H | G | D | FF | Veredicto | Obs |')
A('|---|---|---|---|---|---|---|---|---|---|---|---|')
for g in games:
    A(f"| {g['id']} | {g['nome']} | {g['genres']} | {g['horas']:g} | {g['enc']} | {g['w']} | {g['h']} | {g['g']} | {g['d']} | {g['ff']} | {show(g['v']) if g['v'] in ORD else 'N/A'} | {g['obs'][:70]} |")
A('')
A('---\n')
A('## 🏆 Por Veredicto\n')
for v in reversed(ORD):
    lst=[g for g in rated if g['v']==v]
    if not lst: continue
    A(f'### {show(v)} — {len(lst)} jogo(s)\n')
    A('| Jogo | Horas | Status |')
    A('|---|---|---|')
    for g in sorted(lst, key=lambda x:-x['horas']):
        st='✅' if g['enc']=='S' else ('❌' if g['enc']=='N' else '🔄')
        A(f"| {g['nome']} | {g['horas']:g}h | {st} |")
    A('')
A('---\n')
A('## 📊 Estatísticas\n')
A('### ⭐ Distribuição de Veredictos\n')
A('| Veredicto | Jogos | % |')
A('|---|---|---|')
for v in reversed(ORD):
    if dist[v]: A(f'| {show(v)} | {dist[v]} | {dist[v]/len(rated):.0%} |')
A('')
A('### ⏱️ Top 10 — Mais Horas Jogadas\n')
A('| # | Jogo | Total | Veredicto |')
A('|---|---|---|---|')
for i,g in enumerate(sorted(games,key=lambda x:-x['horas'])[:10],1):
    A(f"| {i} | {g['nome']} | **{g['horas']:g}h** | {show(g['v']) if g['v'] in ORD else 'N/A'} |")
A('')
A('### 🎮 Análise por Gênero Principal\n')
A('| Gênero | Jogos | Horas | Veredicto Médio | Conclusão | Drop% |')
A('|---|---|---|---|---|---|')
gen=defaultdict(list)
for g in games: gen[g['genres'].split(',')[0].strip()].append(g)
for k,v in sorted(gen.items(), key=lambda kv:-len(kv[1])):
    c=sum(1 for x in v if x['enc']=='S'); d=sum(1 for x in v if x['enc']=='N'); fin=max(c+d,1)
    A(f"| {k} | {len(v)} | {sum(x['horas'] for x in v):,.0f}h | {vmed(v)} | {c/fin:.0%} | {d/fin:.0%} |".replace(',','.'))
A('')
A('### 🔥 Dificuldade vs Desempenho\n')
A('| Dificuldade | Jogos | Veredicto Médio | Conclusão |')
A('|---|---|---|---|')
for dv,lbl in [(1,'D=1 Fácil'),(2,'D=2 Médio'),(3,'D=3 Difícil')]:
    v=[g for g in games if g['d']==dv]; fin=[x for x in v if x['enc'] in 'SN']
    c=sum(1 for x in fin if x['enc']=='S')
    A(f'| {lbl} | {len(v)} ({len(v)/total:.0%}) | {vmed(v)} | {c/len(fin):.0%} |')
A('')
A('### 💡 Insights do Jogador\n')
A('| Padrão | Observação |')
A('|---|---|')
A('| **Gameplay é o motor do veredicto** | História amplifica, não sustenta; para souls-likes história não é eixo |')
A('| **Fun Factor 4 → conclusão garantida** | Padrão histórico de 100% |')
A('| **Horas jogadas ≠ qualidade** | Valhalla 131h = Medíocre (jogo gigante + pandemia); Tsushima 65h = Masterpiece |')
A('| **Retentativa de jogo dropado: 0% de sucesso** | Nenhuma passou de Medíocre — diferente de replay de jogo zerado |')
A('| **Walkthrough >=2 = jogo amado (ou DLC)** | Tsushima e Exp33 Masterpiece, TLoU; exceção FFXV (DLCs) |')
A('| **Fórmula Ubisoft = risco alto** | Odyssey/Origins/Outlaws Muito Ruim e drop; exceção parcial: Shadows em NG+ |')
A('| **Souls-like só FromSoftware** | Elden Ring/Sekiro Masterpiece vs LotF/Demon Souls Medíocre, Khazan/Wuchang drop |')
A('| **Survival Horror: 100% conclusão** | Gênero subestimado na coleção |')
A("| **Sem fast travel fácil = drop** | Dragon's Dogma 2, FFXII, Kingdom Come II |")
A('| **Combate simples = drop mesmo com boa escrita** | AC Black Flag Resync: cutscenes boas, dropado em 13h |')
A('')
A('---\n')
A('## 🗓️ Planejamento de Jogos — 2026 (status jul/2026)\n')
A('Disponibilidade: ~12h/semana (~52h/mês; setembro reduzido ~34h por férias 08–27/09).\n')
A('| Jogo | Lançamento | Estimativa | Status |')
A('|---|---|---|---|')
A('| 007 First Light | Jun/26 | 15h | — |')
A('| AC Black Flag Resync | 09/Jul | 45h | ❌ Dropado com 13h (jul/26) — liberou ~32h |')
A('| Beast of Reincarnation | 05/Ago | 45h | Próximo |')
A('| Lord of the Fallen 2 | 10/Ago | 40h | Sugerido p/ Setembro |')
A('| Blood of Dawn | 03/Set | 45h | Outubro |')
A('| Wolverine | 15/Set | 25h | Novembro |')
A('| Phantom Blade | 08/Set | 50h | Nov–Dez |')
A('')
A('**Em andamento (jul/26):** AC Shadows NG+ nightmare (comprado R$60) · Exp33 rezerado (+17h)  ')
A('**FF14:** assinatura pausada; retorno out/nov 2026 antes da expansão Evercold (jan/2027)\n')
A('---\n')
A('## ❌ Jogos Droppados — Motivos\n')
A('| Jogo | Horas | Veredicto | Motivo |')
A('|---|---|---|---|')
for g in sorted(drop,key=lambda x:-x['horas']):
    A(f"| {g['nome']} | {g['horas']:g}h | {show(g['v']) if g['v'] in ORD else 'N/A'} | {g['motivo'] or ''} |")
A('')
A('---\n')
A(f'*Documentação {VERSAO} — atualizada em {HOJE}*  ')
A('*Alterações v11: Ajustes de veredicto confirmados pelo usuário — Valhalla Bom→Medíocre, Demon Souls Bom→Medíocre, Hogwarts Ruim→Bom, Black Flag Resync Muito Ruim→Ruim · Gerador `gerar_doc.py` criado · Spec do agente em `agente_perfil_gamer.md`*  ')
A('*Alterações v10: Sistema de Veredictos (6 níveis, estilo ACG) substitui nota numérica · Componentes 1–4 mantidos internos*  ')
A('*Alterações v9: HLTB e Tempo PS4/PS5 removidos · Exp33 79h W=2 · AC Shadows NG+ R$60 · Black Flag Resync adicionado · Tsushima 65h, HZD 25h*')

open('biblioteca_jogos.md','w',encoding='utf-8').write('\n'.join(L))
print('gerado', VERSAO, '-', len(L), 'linhas |', dict(dist))
